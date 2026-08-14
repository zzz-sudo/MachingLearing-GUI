from __future__ import annotations

import csv
import gzip
import hashlib
import io
import mimetypes
import re
import shutil
import stat
import tarfile
import zipfile
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from typing import Any, Iterable
from uuid import uuid4

from charset_normalizer import from_bytes
from openpyxl import load_workbook

from app.errors import WorkspaceServiceError, import_error
from app.models import AssetRecord, ImportResult, PreviewColumn, TablePreview
from app.storage import WorkspaceStore

PREVIEW_ROW_LIMIT = 100
MAX_ARCHIVE_FILES = 2000
MAX_ARCHIVE_BYTES = 512 * 1024 * 1024
TABLE_SUFFIXES = {".csv", ".xlsx"}
ARCHIVE_SUFFIXES = {".zip", ".tar", ".tgz", ".gz"}


class FileImporter:
    def __init__(self, store: WorkspaceStore) -> None:
        self.store = store

    def import_bytes(
        self,
        project_id: str,
        filename: str,
        content: bytes,
    ) -> ImportResult:
        project = self.store.get_project(project_id)
        safe_name = self._safe_upload_name(filename)
        source_dir = Path(project.path) / "source"
        source_path = self._unique_path(source_dir, safe_name)

        try:
            source_path.write_bytes(content)
        except OSError as error:
            raise import_error(
                "FileAccessError",
                f"无法保存导入文件: {safe_name}",
                "file_import",
                filename=safe_name,
                reason=str(error),
            ) from error

        source_asset = self._record_asset(project_id, Path(project.path), source_path)
        imported_assets = [source_asset]
        preview: TablePreview | None = None
        extracted_count = 0

        suffix = self._compound_suffix(source_path)
        if suffix in ARCHIVE_SUFFIXES:
            extracted_paths = self._extract_archive(source_path, Path(project.path))
            extracted_count = len(extracted_paths)
            for extracted_path in extracted_paths:
                asset = self._record_asset(
                    project_id,
                    Path(project.path),
                    extracted_path,
                    source_asset.id,
                )
                imported_assets.append(asset)
                if preview is None and extracted_path.suffix.lower() in TABLE_SUFFIXES:
                    preview = self._preview_table(extracted_path, asset.id)
        elif source_path.suffix.lower() in TABLE_SUFFIXES:
            preview = self._preview_table(source_path, source_asset.id)
        else:
            raise import_error(
                "UnsupportedFileFormatError",
                f"当前版本不支持该文件格式: {source_path.suffix or safe_name}",
                "file_import",
                filename=safe_name,
                supportedFormats=sorted(TABLE_SUFFIXES | ARCHIVE_SUFFIXES),
            )

        return ImportResult(
            imported_assets=imported_assets,
            preview=preview,
            extracted_count=extracted_count,
        )

    def _record_asset(
        self,
        project_id: str,
        project_root: Path,
        path: Path,
        parent_asset_id: str | None = None,
    ) -> AssetRecord:
        content_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        media_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        return self.store.create_asset(
            project_id=project_id,
            name=path.name,
            relative_path=path.relative_to(project_root).as_posix(),
            media_type=media_type,
            size=path.stat().st_size,
            sha256=content_hash,
            parent_asset_id=parent_asset_id,
        )

    def _preview_table(self, path: Path, asset_id: str) -> TablePreview:
        if path.suffix.lower() == ".csv":
            return self._preview_csv(path, asset_id)
        return self._preview_xlsx(path, asset_id)

    def _preview_csv(self, path: Path, asset_id: str) -> TablePreview:
        raw = path.read_bytes()
        encoding = self._detect_encoding(raw, path.name)
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError as error:
            raise import_error(
                "FileEncodingError",
                f"无法使用检测到的编码读取文件: {path.name}",
                "csv_preview",
                filename=path.name,
                encoding=encoding,
                position=error.start,
            ) from error

        sample = text[:65536]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(io.StringIO(text), dialect)
        rows = list(reader)
        if not rows:
            raise import_error(
                "SpreadsheetParseError",
                f"CSV 文件没有可预览的数据: {path.name}",
                "csv_preview",
                filename=path.name,
            )

        headers = self._normalize_headers(rows[0])
        data_rows = rows[1:]
        preview_rows = [self._row_to_record(headers, row) for row in data_rows[:PREVIEW_ROW_LIMIT]]
        return self._build_preview(
            asset_id=asset_id,
            source_name=path.name,
            format="csv",
            encoding=encoding,
            sheet_name=None,
            headers=headers,
            rows=preview_rows,
            row_count=len(data_rows),
        )

    def _preview_xlsx(self, path: Path, asset_id: str) -> TablePreview:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            iterator = worksheet.iter_rows(values_only=True)
            first_row = next(iterator, None)
            if first_row is None:
                raise import_error(
                    "SpreadsheetParseError",
                    f"工作簿没有可预览的数据: {path.name}",
                    "xlsx_preview",
                    filename=path.name,
                )
            headers = self._normalize_headers(first_row)
            preview_rows = []
            for row in iterator:
                if len(preview_rows) >= PREVIEW_ROW_LIMIT:
                    break
                preview_rows.append(self._row_to_record(headers, row))
            row_count = max((worksheet.max_row or 1) - 1, 0)
            sheet_name = worksheet.title
            workbook.close()
        except WorkspaceServiceError:
            raise
        except (OSError, ValueError, KeyError, zipfile.BadZipFile) as error:
            raise import_error(
                "SpreadsheetParseError",
                f"无法读取 XLSX 工作簿: {path.name}",
                "xlsx_preview",
                filename=path.name,
                reason=str(error),
            ) from error

        return self._build_preview(
            asset_id=asset_id,
            source_name=path.name,
            format="xlsx",
            encoding=None,
            sheet_name=sheet_name,
            headers=headers,
            rows=preview_rows,
            row_count=row_count,
        )

    def _extract_archive(self, archive_path: Path, project_root: Path) -> list[Path]:
        destination = project_root / "extracted" / f"{archive_path.stem}-{uuid4().hex[:8]}"
        destination.mkdir(parents=True, exist_ok=False)
        suffix = self._compound_suffix(archive_path)
        try:
            if suffix == ".zip":
                return self._extract_zip(archive_path, destination)
            if suffix in {".tar", ".tgz"}:
                return self._extract_tar(archive_path, destination)
            return self._extract_gzip(archive_path, destination)
        except WorkspaceServiceError:
            raise
        except (OSError, EOFError, gzip.BadGzipFile, tarfile.TarError, zipfile.BadZipFile) as error:
            raise import_error(
                "ArchiveExtractionError",
                f"无法解压文件: {archive_path.name}",
                "archive_extract",
                filename=archive_path.name,
                reason=str(error),
            ) from error

    def _extract_zip(self, archive_path: Path, destination: Path) -> list[Path]:
        extracted: list[Path] = []
        total_size = 0
        with zipfile.ZipFile(archive_path) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ARCHIVE_FILES:
                self._archive_limit_error(archive_path.name, "文件数量超过限制")
            for info in infos:
                name = self._decode_zip_name(info)
                if info.is_dir():
                    continue
                if stat.S_ISLNK(info.external_attr >> 16):
                    self._archive_limit_error(archive_path.name, "压缩包包含符号链接")
                total_size += info.file_size
                if total_size > MAX_ARCHIVE_BYTES:
                    self._archive_limit_error(archive_path.name, "解压后体积超过限制")
                output_path = self._safe_archive_path(destination, name)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as source, output_path.open("wb") as target:
                    shutil.copyfileobj(source, target)
                extracted.append(output_path)
        return extracted

    def _extract_tar(self, archive_path: Path, destination: Path) -> list[Path]:
        extracted: list[Path] = []
        total_size = 0
        with tarfile.open(archive_path, mode="r:*") as archive:
            members = archive.getmembers()
            if len(members) > MAX_ARCHIVE_FILES:
                self._archive_limit_error(archive_path.name, "文件数量超过限制")
            for member in members:
                if member.isdir():
                    continue
                if not member.isfile():
                    self._archive_limit_error(archive_path.name, "压缩包包含非普通文件")
                total_size += member.size
                if total_size > MAX_ARCHIVE_BYTES:
                    self._archive_limit_error(archive_path.name, "解压后体积超过限制")
                output_path = self._safe_archive_path(destination, member.name)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                source = archive.extractfile(member)
                if source is None:
                    continue
                with source, output_path.open("wb") as target:
                    shutil.copyfileobj(source, target)
                extracted.append(output_path)
        return extracted

    def _extract_gzip(self, archive_path: Path, destination: Path) -> list[Path]:
        output_name = archive_path.stem or "decompressed-file"
        output_path = self._safe_archive_path(destination, output_name)
        with gzip.open(archive_path, "rb") as source, output_path.open("wb") as target:
            written = 0
            while chunk := source.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_ARCHIVE_BYTES:
                    self._archive_limit_error(archive_path.name, "解压后体积超过限制")
                target.write(chunk)
        return [output_path]

    @staticmethod
    def _detect_encoding(raw: bytes, filename: str) -> str:
        if raw.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
        if raw.startswith(b"\xff\xfe"):
            return "utf-16-le"
        if raw.startswith(b"\xfe\xff"):
            return "utf-16-be"
        try:
            raw.decode("utf-8", errors="strict")
            return "utf-8"
        except UnicodeDecodeError:
            pass
        try:
            raw.decode("gb18030", errors="strict")
            return "gb18030"
        except UnicodeDecodeError:
            match = from_bytes(raw).best()
            if match is None or match.encoding is None:
                raise import_error(
                    "FileEncodingError",
                    f"无法可靠识别文本编码: {filename}",
                    "csv_encoding_detect",
                    filename=filename,
                    attemptedEncodings=["utf-8", "gb18030"],
                )
            return match.encoding

    @staticmethod
    def _normalize_headers(values: Iterable[Any]) -> list[str]:
        headers: list[str] = []
        used: dict[str, int] = {}
        for index, value in enumerate(values, start=1):
            base = str(value).strip() if value is not None else ""
            base = base or f"未命名字段_{index}"
            count = used.get(base, 0) + 1
            used[base] = count
            headers.append(base if count == 1 else f"{base}_{count}")
        return headers

    @staticmethod
    def _row_to_record(headers: list[str], values: Iterable[Any]) -> dict[str, Any]:
        row = list(values)
        return {
            header: FileImporter._json_value(row[index] if index < len(row) else None)
            for index, header in enumerate(headers)
        }

    @staticmethod
    def _json_value(value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return value

    @staticmethod
    def _build_preview(
        asset_id: str,
        source_name: str,
        format: str,
        encoding: str | None,
        sheet_name: str | None,
        headers: list[str],
        rows: list[dict[str, Any]],
        row_count: int,
    ) -> TablePreview:
        columns = []
        for header in headers:
            values = [row.get(header) for row in rows]
            columns.append(
                PreviewColumn(
                    name=header,
                    inferred_type=FileImporter._infer_type(values),
                    null_count=sum(value in (None, "") for value in values),
                )
            )
        return TablePreview(
            asset_id=asset_id,
            source_name=source_name,
            format=format,
            encoding=encoding,
            sheet_name=sheet_name,
            row_count=row_count,
            column_count=len(headers),
            columns=columns,
            rows=rows,
        )

    @staticmethod
    def _infer_type(values: list[Any]) -> str:
        present = [value for value in values if value not in (None, "")]
        if not present:
            return "empty"
        if all(isinstance(value, bool) for value in present):
            return "boolean"
        if all(isinstance(value, int) and not isinstance(value, bool) for value in present):
            return "integer"
        if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
            return "number"
        text_values = [value.strip() for value in present if isinstance(value, str)]
        if len(text_values) == len(present):
            if all(re.fullmatch(r"[+-]?\d+", value) for value in text_values):
                return "integer"
            if all(
                re.fullmatch(r"[+-]?(?:\d+\.\d*|\d*\.\d+|\d+)", value)
                for value in text_values
            ):
                return "number"
        return "text"

    @staticmethod
    def _safe_upload_name(filename: str) -> str:
        name = Path(filename.replace("\\", "/")).name.strip()
        if not name or name in {".", ".."}:
            raise import_error(
                "FileAccessError",
                "导入文件名无效",
                "file_import",
                filename=filename,
            )
        return name

    @staticmethod
    def _unique_path(directory: Path, filename: str) -> Path:
        candidate = directory / filename
        if not candidate.exists():
            return candidate
        return directory / f"{candidate.stem}-{uuid4().hex[:8]}{candidate.suffix}"

    @staticmethod
    def _compound_suffix(path: Path) -> str:
        lower_name = path.name.lower()
        if lower_name.endswith(".tar.gz"):
            return ".tgz"
        return path.suffix.lower()

    @staticmethod
    def _safe_archive_path(destination: Path, member_name: str) -> Path:
        pure_path = PurePosixPath(member_name.replace("\\", "/"))
        if pure_path.is_absolute() or ".." in pure_path.parts:
            raise import_error(
                "ArchiveExtractionError",
                "压缩包包含非法路径",
                "archive_extract",
                memberName=member_name,
            )
        output_path = (destination / Path(*pure_path.parts)).resolve()
        if destination.resolve() not in output_path.parents:
            raise import_error(
                "ArchiveExtractionError",
                "压缩包文件超出项目目录",
                "archive_extract",
                memberName=member_name,
            )
        return output_path

    @staticmethod
    def _decode_zip_name(info: zipfile.ZipInfo) -> str:
        if info.flag_bits & 0x800:
            return info.filename
        try:
            return info.filename.encode("cp437").decode("gb18030")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return info.filename

    @staticmethod
    def _archive_limit_error(filename: str, reason: str) -> None:
        raise import_error(
            "ArchiveExtractionError",
            f"压缩文件不符合解压限制: {filename}",
            "archive_extract",
            filename=filename,
            reason=reason,
        )
