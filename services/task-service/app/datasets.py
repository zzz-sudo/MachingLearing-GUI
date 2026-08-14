from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.parquet as parquet
from openpyxl import load_workbook

from app.errors import import_error
from app.models import DatasetColumnSpec, DatasetCreate, DatasetVersion
from app.storage import WorkspaceStore

SUPPORTED_TYPES = {"text", "integer", "number", "boolean"}


class DatasetService:
    def __init__(self, store: WorkspaceStore) -> None:
        self.store = store

    def create(self, project_id: str, payload: DatasetCreate) -> DatasetVersion:
        project = self.store.get_project(project_id)
        asset = self.store.get_asset(payload.asset_id)
        preview = self.store.get_preview(asset.id)
        expected = {column.name for column in preview.columns}
        supplied = {column.name for column in payload.columns}
        if asset.project_id != project_id or expected != supplied or any(column.data_type not in SUPPORTED_TYPES for column in payload.columns):
            raise import_error(
                "DatasetSchemaError", "字段确认与源表结构不一致", "dataset_create",
                expectedColumns=sorted(expected), suppliedColumns=sorted(supplied),
            )

        source_path = Path(project.path) / asset.relative_path
        records = self._read_records(source_path, preview.encoding, preview.sheet_name)
        converted = [self._convert_record(record, payload.columns, index + 2) for index, record in enumerate(records)]
        schema = pa.schema([pa.field(column.name, self._arrow_type(column.data_type)) for column in payload.columns])
        table = pa.Table.from_pylist(converted, schema=schema)
        existing = [item for item in self.store.list_dataset_versions(project_id) if item.source_asset_id == asset.id]
        output_path = Path(project.path) / "datasets" / f"{Path(asset.name).stem}-v{len(existing) + 1}.parquet"
        parquet.write_table(table, output_path, compression="zstd")
        return self.store.create_dataset_version(
            project_id, asset.id, output_path.relative_to(Path(project.path)).as_posix(),
            table.num_rows, payload.columns,
        )

    @staticmethod
    def _read_records(path: Path, encoding: str | None, sheet_name: str | None) -> list[dict[str, Any]]:
        if path.suffix.lower() == ".csv":
            text = path.read_bytes().decode(encoding or "utf-8")
            return list(csv.DictReader(io.StringIO(text)))
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(rows)]
        records = [dict(zip(headers, row, strict=False)) for row in rows]
        workbook.close()
        return records

    @staticmethod
    def _convert_record(record: dict[str, Any], columns: list[DatasetColumnSpec], row_number: int) -> dict[str, Any]:
        converted: dict[str, Any] = {}
        for column in columns:
            value = record.get(column.name)
            if value in (None, ""):
                converted[column.name] = None
                continue
            try:
                if column.data_type == "integer":
                    converted[column.name] = int(value)
                elif column.data_type == "number":
                    converted[column.name] = float(value)
                elif column.data_type == "boolean":
                    converted[column.name] = value if isinstance(value, bool) else str(value).lower() in {"true", "1", "yes", "是"}
                else:
                    converted[column.name] = str(value)
            except (TypeError, ValueError) as error:
                raise import_error(
                    "DatasetSchemaError", f"第 {row_number} 行字段转换失败: {column.name}",
                    "dataset_create", row=row_number, column=column.name,
                    dataType=column.data_type, value=str(value),
                ) from error
        return converted

    @staticmethod
    def _arrow_type(data_type: str) -> pa.DataType:
        return {"text": pa.string(), "integer": pa.int64(), "number": pa.float64(), "boolean": pa.bool_()}[data_type]
